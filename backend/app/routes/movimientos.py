from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import desc
from datetime import date, timedelta
from typing import Optional
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models.movimiento import Movimiento, TipoMovimiento
from app.models.insumo import Insumo
from app.models.usuario import Usuario
from app.models.sala import Sala
from app.schemas.movimiento import MovimientoCreate, MovimientoResponse, MovimientoEnriquecido
from app.schemas.comun import PaginatedResponse
from app.utils.deps import get_usuario_actual, require_operador

router = APIRouter(prefix="/movimientos", tags=["Movimientos"])


def _enriquecer(m: Movimiento) -> MovimientoEnriquecido:
    return MovimientoEnriquecido(
        id=m.id,
        tipo=m.tipo,
        cantidad=m.cantidad,
        motivo=m.motivo,
        fecha=m.fecha,
        insumo=m.insumo.nombre if m.insumo else "Desconocido",
        sala=m.insumo.sala.nombre if m.insumo and m.insumo.sala else None,
        usuario=m.usuario.nombre if m.usuario else "Desconocido"
    )


def _build_query(
    db: Session,
    insumo: Optional[str] = None,
    tipo: Optional[str] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
):
    """Query base con filtros opcionales para listado y exportacion.

    Para filtrar por nombre de insumo usamos una subquery sobre la tabla
    insumos en vez de un JOIN explicito, para no interferir con los
    joinedload que cargan las relaciones para _enriquecer().

    fecha_hasta es inclusivo: se suma un dia para incluir todo el dia UTC.
    """
    q = (
        db.query(Movimiento)
        .options(
            joinedload(Movimiento.insumo).joinedload(Insumo.sala),
            joinedload(Movimiento.usuario),
        )
        .order_by(desc(Movimiento.fecha))
    )
    if insumo:
        ids = (
            db.query(Insumo.id)
            .filter(Insumo.nombre.ilike(f"%{insumo}%"))
            .subquery()
        )
        q = q.filter(Movimiento.insumo_id.in_(ids))
    if tipo and tipo in ("entrada", "salida"):
        q = q.filter(Movimiento.tipo == tipo)
    if fecha_desde:
        q = q.filter(Movimiento.fecha >= fecha_desde)
    if fecha_hasta:
        # Sumar 1 dia para que fecha_hasta sea inclusivo
        q = q.filter(Movimiento.fecha < fecha_hasta + timedelta(days=1))
    return q


# ---------------------------------------------------------------------------
# IMPORTANTE: /exportar debe ir ANTES de las rutas dinamicas /{id},
# de lo contrario FastAPI lo interpreta como un entero y devuelve 422.
# ---------------------------------------------------------------------------

@router.get("/exportar")
def exportar_movimientos(
    formato: str = "csv",
    insumo: Optional[str] = None,
    tipo: Optional[str] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(get_usuario_actual),
):
    """Exporta movimientos como CSV o Excel con los mismos filtros del listado.

    formato=csv  -> archivo .csv con BOM UTF-8 (abre bien en Excel).
    formato=xlsx -> archivo .xlsx con cabecera coloreada y columnas autoajustadas.
    """
    movimientos = _build_query(db, insumo, tipo, fecha_desde, fecha_hasta).all()
    filas = [
        [
            m.tipo.value if hasattr(m.tipo, "value") else m.tipo,
            m.insumo.nombre if m.insumo else "Desconocido",
            m.insumo.sala.nombre if m.insumo and m.insumo.sala else "",
            m.cantidad,
            m.motivo or "",
            m.fecha.strftime("%d/%m/%Y %H:%M") if m.fecha else "",
            m.usuario.nombre if m.usuario else "Desconocido",
        ]
        for m in movimientos
    ]
    cabecera = ["Tipo", "Insumo", "Sala", "Cantidad", "Motivo", "Fecha", "Usuario"]

    if formato == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"

        # Cabecera con estilo
        header_fill = PatternFill("solid", fgColor="0F766E")  # teal-700
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx, titulo in enumerate(cabecera, 1):
            cell = ws.cell(row=1, column=col_idx, value=titulo)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Datos
        for fila in filas:
            ws.append(fila)

        # Autoajustar ancho de columnas
        for col_idx, _ in enumerate(cabecera, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or ""))
                 for r in range(1, ws.max_row + 1)),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=movimientos_hestia.xlsx"},
        )

    # Por defecto: CSV con BOM UTF-8 para compatibilidad con Excel
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(cabecera)
    writer.writerows(filas)
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=movimientos_hestia.csv"},
    )


@router.get("/", response_model=PaginatedResponse[MovimientoEnriquecido])
def listar_movimientos(
    skip: int = 0,
    limit: int = 20,
    insumo: Optional[str] = None,
    tipo: Optional[str] = None,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(get_usuario_actual),
):
    """Lista movimientos con filtros opcionales: insumo (texto), tipo,
    fecha_desde y fecha_hasta. Todos los filtros son acumulables.
    """
    q = _build_query(db, insumo, tipo, fecha_desde, fecha_hasta)
    total = q.count()
    movimientos = q.offset(skip).limit(limit).all()
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "data": [_enriquecer(m) for m in movimientos],
    }


@router.post("/", response_model=MovimientoResponse)
def registrar_movimiento(
    mov: MovimientoCreate,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(require_operador),
):
    insumo = db.query(Insumo).filter(Insumo.id == mov.insumo_id).with_for_update().first()
    if not insumo:
        raise HTTPException(status_code=404, detail="Insumo no encontrado")

    if not insumo.activo:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                f"'{insumo.nombre}' esta inactivo. Reactivalo antes de "
                "registrar movimientos."
            ),
        )

    if mov.tipo == TipoMovimiento.salida:
        if insumo.stock_actual < mov.cantidad:
            raise HTTPException(
                status_code=400,
                detail=f"Stock insuficiente. Disponible: {insumo.stock_actual}",
            )
        insumo.stock_actual -= mov.cantidad
    else:
        insumo.stock_actual += mov.cantidad

    nuevo_mov = Movimiento(**mov.model_dump(), usuario_id=usuario.id)
    db.add(nuevo_mov)
    db.commit()
    db.refresh(nuevo_mov)
    return nuevo_mov


@router.get("/insumo/{insumo_id}", response_model=PaginatedResponse[MovimientoEnriquecido])
def historial_por_insumo(
    insumo_id: int,
    skip: int = 0,
    limit: int = 20,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(get_usuario_actual),
):
    if not db.query(Insumo).filter(Insumo.id == insumo_id).first():
        raise HTTPException(status_code=404, detail="Insumo no encontrado")
    total = db.query(Movimiento).filter(Movimiento.insumo_id == insumo_id).count()
    movimientos = (
        _build_query(db)
        .filter(Movimiento.insumo_id == insumo_id)
        .offset(skip).limit(limit).all()
    )
    return {"total": total, "skip": skip, "limit": limit,
            "data": [_enriquecer(m) for m in movimientos]}


@router.get("/sala/{sala_id}", response_model=PaginatedResponse[MovimientoEnriquecido])
def historial_por_sala(
    sala_id: int,
    skip: int = 0,
    limit: int = 20,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(get_usuario_actual),
):
    if not db.query(Sala).filter(Sala.id == sala_id).first():
        raise HTTPException(status_code=404, detail="Sala no encontrada")
    total = (
        db.query(Movimiento)
        .join(Insumo, Movimiento.insumo_id == Insumo.id)
        .filter(Insumo.sala_id == sala_id)
        .count()
    )
    movimientos = (
        _build_query(db)
        .join(Insumo, Movimiento.insumo_id == Insumo.id)
        .filter(Insumo.sala_id == sala_id)
        .offset(skip).limit(limit).all()
    )
    return {"total": total, "skip": skip, "limit": limit,
            "data": [_enriquecer(m) for m in movimientos]}
